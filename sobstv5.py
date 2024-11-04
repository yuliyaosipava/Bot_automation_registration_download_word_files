import os
from openpyxl import load_workbook, Workbook

TOTAL_INCOME_LABELS = ['ВСЕГО', 'ВСЕГО доходов']
GRANTS_LABELS = ['БЕЗВОЗМЕЗДНЫЕ ПОСТУПЛЕНИЯ', 'Безвозмездные поступления']
EXECUTED_COLUMN = 'Исполнено'  # Ensure this is correct

def extract_values(file_path):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        total_income = None
        grants = None
        headers = {cell.value: cell.column_letter for cell in next(sheet.iter_rows(min_row=1, max_row=1))}
        executed_col = headers.get(EXECUTED_COLUMN, None)
        
        found_total_income = False
        found_grants = False

        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            if not found_total_income and row[0].value and str(row[0].value).strip().lower() in [label.lower() for label in TOTAL_INCOME_LABELS]:
                total_income = sheet[f'{executed_col}{row[0].row}'].value
                found_total_income = True  # Set flag to True after first encounter
            
            if not found_grants and row[0].value and str(row[0].value).strip().lower() in [label.lower() for label in GRANTS_LABELS]:
                grants = sheet[f'{executed_col}{row[0].row}'].value
                found_grants = True  # Set flag to True after first encounter
            
            if found_total_income and found_grants:
                break  # Exit loop once both values are found
        
        return total_income, grants
    except Exception as e:
        print(f'Error processing file {file_path}: {e}')
        return None, None

def process_files_in_directory(input_dir, output_file):
    results = []
    for filename in os.listdir(input_dir):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(input_dir, filename)
            total_income, grants = extract_values(file_path)
            if total_income is not None and grants is not None:
                try:
                    total_income = float(str(total_income).replace(" ", "").replace(",", "."))
                    grants = float(str(grants).replace(" ", "").replace(",", "."))
                    if total_income != 0:  # Check if total_income is zero
                        difference = total_income - grants
                        difference_percentage = (difference / total_income) * 100 if total_income != 0 else "N/A"
                        results.append([filename, total_income, grants, difference, difference_percentage])
                        print(f'File: {filename}')
                        print(f'TOTAL INCOME: {total_income}')
                        print(f'GRANTS: {grants}')
                        print(f'Difference: {difference_percentage:.2f}%')
                    else:
                        results.append([filename, total_income, grants, "N/A", "N/A"])
                        print(f'Skipping percentage calculation for file {filename} due to zero total income')
                except ValueError:
                    print(f'Error converting values in file: {filename}')
            else:
                print(f'Necessary values not found in file: {filename}')
    save_results_to_excel(results, output_file)

def save_results_to_excel(results, output_file):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["File", "TOTAL INCOME", "GRANTS", "Difference", "Difference (%)"])
    for result in results:
        sheet.append(result)
    workbook.save(output_file)

input_dir = 'c:\\Users\\user\\Documents\\doci\\xls1\\'
output_file = 'c:\\Users\\user\\Documents\\doci\\result\\output6.xlsx'
process_files_in_directory(input_dir, output_file)
